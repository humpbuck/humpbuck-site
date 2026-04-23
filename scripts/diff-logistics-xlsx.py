# -*- coding: utf-8 -*-
"""Compare lib/data/logistics-rates.json with source XLSX workbooks; report gaps & mismatches."""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
JSON_PATH = ROOT / "lib" / "data" / "logistics-rates.json"
CAINIAO = Path(r"c:\Users\Administrator\Desktop\《菜鸟跨境产品报价表-20260418010526》.xlsx")
YANWEN = Path(r"c:\Users\Administrator\Desktop\东莞燕文报价单20260413版.xlsx")


def parse_weight_band(s: str) -> tuple[float, float] | None:
    if not s or not isinstance(s, str):
        return None
    s = s.strip().replace("＜", "<").replace("≤", "<=").replace(" ", "")
    m = re.match(r"^([\d.]+)<W<=([\d.]+)$", s)
    if m:
        return float(m.group(1)), float(m.group(2))
    m = re.match(r"^([\d.]+)<W≤([\d.]+)$", s)
    if m:
        return float(m.group(1)), float(m.group(2))
    return None


def load_cainiao_sheet(wb, sheet_name: str) -> dict[str, list[dict]]:
    ws = wb[sheet_name]
    by_country: dict[str, list[dict]] = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        country = row[1]
        band_s = row[4]
        rmb_kg = row[5]
        reg = row[6]
        if not country or not band_s:
            continue
        if not isinstance(country, str):
            continue
        band = parse_weight_band(str(band_s))
        if not band:
            continue
        try:
            pk = float(rmb_kg)
            pr = float(reg)
        except (TypeError, ValueError):
            continue
        c = country.strip()
        by_country.setdefault(c, []).append(
            {"wMin": band[0], "wMax": band[1], "rmbPerKg": pk, "regRmb": pr}
        )
    for c in by_country:
        by_country[c].sort(key=lambda x: (x["wMin"], x["wMax"]))
    return by_country


def band_key(b: dict) -> tuple:
    return (
        round(float(b["wMin"]), 6),
        round(float(b["wMax"]), 6),
        round(float(b["rmbPerKg"]), 4),
        round(float(b["regRmb"]), 4),
    )


def yw_band_key(b: dict) -> tuple:
    return (
        round(float(b.get("wMin") or 0), 6),
        round(float(b.get("wMax") or 0), 6),
        round(float(b["rmbPerKg"]), 4),
        round(float(b["pieceRmb"]), 4),
        round(float(b.get("minChargeKg") or 0), 6),
    )


def compare_cainiao_side(label: str, xlsx_map: dict, json_map: dict) -> None:
    x_keys = set(xlsx_map)
    j_keys = set(json_map)
    only_x = sorted(x_keys - j_keys)
    only_j = sorted(j_keys - x_keys)
    if only_x:
        print(f"\n[{label}] countries in XLSX but missing in JSON ({len(only_x)}):")
        for k in only_x[:80]:
            print(f"  + {k}")
        if len(only_x) > 80:
            print(f"  ... +{len(only_x) - 80} more")
    if only_j:
        print(f"\n[{label}] countries in JSON but not in XLSX ({len(only_j)}):")
        for k in only_j[:40]:
            print(f"  - {k}")
        if len(only_j) > 40:
            print(f"  ... +{len(only_j) - 40} more")

    mismatches = 0
    for c in sorted(x_keys & j_keys):
        xb = [band_key(x) for x in xlsx_map[c]]
        jb = [band_key(x) for x in json_map[c]]
        if xb != jb:
            mismatches += 1
            print(f"\n[{label}] band mismatch: {c}")
            print(f"  xlsx ({len(xb)}): {xb[:6]}{'...' if len(xb) > 6 else ''}")
            print(f"  json ({len(jb)}): {jb[:6]}{'...' if len(jb) > 6 else ''}")
    if mismatches == 0 and not only_x and not only_j:
        print(f"\n[{label}] OK — same country set and identical bands.")


def load_yanwen_484_resolved(
    wb, name_to_iso: dict[str, str]
) -> tuple[dict[str, list[dict] | dict[str, list[dict]]], list[str]]:
    """Match export-logistics-rates: CA/AU nested zones 1–4; others flat."""
    ws = wb["燕文专线追踪-特货"]
    by_iso: dict[str, list[dict] | dict[str, list[dict]]] = {}
    skipped: list[str] = []
    zone_split_countries = {"加拿大", "澳大利亚"}

    def append_band(iso: str, band: dict, zone_digit: str | None) -> None:
        if zone_digit is not None:
            cur = by_iso.get(iso)
            if isinstance(cur, list):
                return
            if not isinstance(cur, dict):
                by_iso[iso] = {}
                cur = by_iso[iso]
            assert isinstance(cur, dict)
            cur.setdefault(zone_digit, []).append(band)
            return
        cur = by_iso.get(iso)
        if isinstance(cur, dict):
            return
        if cur is None:
            by_iso[iso] = []
        assert isinstance(by_iso[iso], list)
        by_iso[iso].append(band)

    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or row[0] is None:
            continue
        name = row[1]
        code = row[2]
        if not isinstance(name, str) or name.strip() in ("国家", "Country"):
            continue
        name_k = name.strip()
        z = str(code).strip() if code is not None else ""
        zone_digit: str | None = None
        if name_k in zone_split_countries:
            m = re.match(r"^([1-4])区$", z)
            if not m:
                continue
            zone_digit = m.group(1)
        iso = None
        if isinstance(code, str):
            cc = code.strip().upper()
            if len(cc) == 2 and cc.isalpha():
                iso = cc
        if iso is None:
            iso = name_to_iso.get(name_k)
        if iso is None:
            skipped.append(f"{name_k} (code={code!r})")
            continue
        try:
            pk = float(row[3])
            fee = float(row[4])
        except (TypeError, ValueError):
            skipped.append(f"{name_k} (non-numeric fee row)")
            continue
        band_s = row[5]
        min_w = row[6]
        wmin, wmax = None, None
        if isinstance(band_s, str) and " - " in band_s:
            a, b = band_s.split(" - ", 1)
            try:
                wmin = float(a.strip())
                wmax = float(b.strip())
            except ValueError:
                pass
        if wmin is None:
            skipped.append(f"{name_k} (bad band {band_s!r})")
            continue
        min_charge = None
        if min_w is not None:
            try:
                min_charge = float(min_w)
            except (TypeError, ValueError):
                pass
        append_band(
            iso,
            {
                "wMin": wmin,
                "wMax": wmax,
                "rmbPerKg": pk,
                "pieceRmb": fee,
                "minChargeKg": min_charge,
            },
            zone_digit,
        )

    for cc, val in list(by_iso.items()):
        if isinstance(val, list):
            val.sort(key=lambda x: (x["wMin"] or 0, x["wMax"] or 0))
        elif isinstance(val, dict):
            for zd in val:
                val[zd].sort(key=lambda x: (x["wMin"] or 0, x["wMax"] or 0))
    return by_iso, skipped


def _yanwen_bands_signature(entry: Any) -> Any:
    if isinstance(entry, list):
        return [yw_band_key(x) for x in entry]
    if isinstance(entry, dict):
        return {k: [yw_band_key(x) for x in v] for k, v in sorted(entry.items())}
    return None


def compare_yanwen(xlsx_map: dict, json_map: dict, skipped: list[str]) -> None:
    if skipped:
        uniq = sorted(set(skipped))
        print(f"\n[Yanwen 484] rows skipped during XLSX parse ({len(uniq)} unique):")
        for s in uniq[:30]:
            print(f"  ? {s}")
        if len(uniq) > 30:
            print(f"  ... +{len(uniq) - 30} more")

    x_keys = set(xlsx_map)
    j_keys = set(json_map)
    only_x = sorted(x_keys - j_keys)
    only_j = sorted(j_keys - x_keys)
    if only_x:
        print(f"\n[Yanwen 484] ISO in XLSX (resolved) missing in JSON ({len(only_x)}):")
        for k in only_x:
            print(f"  + {k}")
    if only_j:
        print(f"\n[Yanwen 484] ISO in JSON not produced from XLSX ({len(only_j)}):")
        for k in only_j:
            print(f"  - {k}")

    mismatches = 0
    for cc in sorted(x_keys & j_keys):
        xs = _yanwen_bands_signature(xlsx_map[cc])
        js = _yanwen_bands_signature(json_map[cc])
        if xs != js:
            mismatches += 1
            print(f"\n[Yanwen 484] band mismatch: {cc}")
            print(f"  xlsx: {xs}")
            print(f"  json: {js}")
    if mismatches == 0 and not only_x and not only_j:
        print("\n[Yanwen 484] OK — same ISO set and identical bands.")


def main() -> None:
    data = json.loads(JSON_PATH.read_text(encoding="utf-8"))
    name_to_iso: dict[str, str] = dict(data.get("cainiaoCountryToIso") or {})

    if not CAINIAO.is_file():
        print("Missing Cainiao file:", CAINIAO, file=sys.stderr)
    if not YANWEN.is_file():
        print("Missing Yanwen file:", YANWEN, file=sys.stderr)

    print("=== Logistics rates diff (JSON vs Desktop XLSX) ===")
    print("JSON:", JSON_PATH)

    if CAINIAO.is_file():
        wb = load_workbook(CAINIAO, read_only=True, data_only=True)
        for sheet_key, json_key in (
            ("菜鸟轻小件-带电（S5059）", "S5059"),
            ("菜鸟标准挂号-带电（OH）", "OH"),
        ):
            xlsx_map = load_cainiao_sheet(wb, sheet_key)
            json_map = (data.get("cainiao") or {}).get(json_key) or {}
            compare_cainiao_side(f"Cainiao {json_key}", xlsx_map, json_map)
        wb.close()

    if YANWEN.is_file():
        wb = load_workbook(YANWEN, read_only=True, data_only=True)
        xlsx_y, skipped = load_yanwen_484_resolved(wb, name_to_iso)
        wb.close()
        json_y = data.get("yanwen484") or {}
        compare_yanwen(xlsx_y, json_y, skipped)

    fb = data.get("cainiaoIsoFallback")
    if fb:
        print(f"\n[Note] JSON keeps cainiaoIsoFallback for {len(fb)} ISO2 keys (not from XLSX).")


if __name__ == "__main__":
    main()
