# -*- coding: utf-8 -*-
"""
Export Yanwen postcode → lane zone (1–4) from Desktop XLSX into lib/data/yanwen-postcode-zones.json.

Sources (sheet names must match workbook):
  - 国家分组-澳洲邮编  → AU numeric postcodes 0–9999
  - 加拿大专线产品邮编分区 → CA FSA / full-code prefixes → zone

Cainiao bundled XLSX has no AU postcode grid; product prices use 澳大利亚/N区 — we document that
AU zones align with this Yanwen extract until a separate Cainiao grid is added.
"""
from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "lib" / "data" / "yanwen-postcode-zones.json"
YANWEN = Path(r"c:\Users\Administrator\Desktop\东莞燕文报价单20260413版.xlsx")


def zone_digit_from_cell(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    m = re.match(r"^([1-4])区", s)
    return m.group(1) if m else None


def parse_au_tokens_to_intervals(text: str) -> list[tuple[int, int]]:
    """Parse '1000-1935, 2000, 3063' into [(lo,hi), ...]."""
    out: list[tuple[int, int]] = []
    for raw in re.split(r"[,，]", text):
        t = raw.strip()
        if not t:
            continue
        t = re.sub(r"\s+", "", t)
        m = re.match(r"^(\d+)\s*-\s*(\d+)$", t)
        if m:
            lo, hi = int(m.group(1)), int(m.group(2))
            if lo > hi:
                lo, hi = hi, lo
            out.append((lo, hi))
            continue
        if t.isdigit():
            n = int(t)
            out.append((n, n))
    return out


def au_apply_intervals(zones: list[int], lo: int, hi: int, z: int) -> None:
    """Mark zones[canonical_pc] = z for pc in 0..9999."""
    for n in range(lo, hi + 1):
        if 0 <= n <= 9999:
            if zones[n] == 0:
                zones[n] = z


def compress_au(zones: list[int]) -> list[dict]:
    out: list[dict] = []
    i = 0
    while i < 10000:
        if zones[i] == 0:
            i += 1
            continue
        z = zones[i]
        j = i
        while j < 10000 and zones[j] == z:
            j += 1
        out.append({"lo": i, "hi": j - 1, "zone": z})
        i = j
    return out


def load_au_zones(wb) -> list[dict]:
    ws = wb["国家分组-澳洲邮编"]
    zones = [0] * 10000
    current: str | None = None

    for row in ws.iter_rows(values_only=True):
        a, b, c, d, e = (row + (None,) * 5)[:5]
        zd = zone_digit_from_cell(a)
        if zd:
            current = zd
        if b is not None and current:
            if isinstance(b, str):
                for lo, hi in parse_au_tokens_to_intervals(b):
                    au_apply_intervals(zones, lo, hi, int(current))
            elif isinstance(b, (int, float)):
                n = int(b)
                if 0 <= n <= 9999:
                    zones[n] = int(current)
        zd_d = zone_digit_from_cell(d)
        if zd_d and e is not None:
            if isinstance(e, (int, float)):
                n = int(e)
                if 0 <= n <= 9999:
                    zones[n] = int(zd_d)

    return compress_au(zones)


def normalize_ca_token(s: str) -> str:
    return re.sub(r"[\s-]+", "", s.upper())


def zone_from_ca_cell(val) -> int | None:
    if val is None:
        return None
    s = str(val).strip()
    m = re.match(r"^([1-4])区", s)
    if m:
        return int(m.group(1))
    if s in ("一区",):
        return 1
    if s in ("二区",):
        return 2
    if s in ("三区",):
        return 3
    if s in ("四区",):
        return 4
    return None


def load_ca_fsa_map(wb) -> dict[str, int]:
    ws = wb["加拿大专线产品邮编分区"]
    best: dict[str, int] = {}
    for row in ws.iter_rows(min_row=5, values_only=True):
        pairs = [(row[i], row[i + 1]) for i in range(0, min(len(row), 6), 2) if i + 1 < len(row)]
        for code_cell, zone_cell in pairs:
            if code_cell is None or zone_cell is None:
                continue
            z = zone_from_ca_cell(zone_cell)
            if z is None:
                continue
            code = normalize_ca_token(str(code_cell))
            if len(code) < 3:
                continue
            fsa3 = code[:3]
            for key in {code, fsa3}:
                if key not in best:
                    best[key] = z
                elif best[key] != z:
                    pass
    return best


def main() -> None:
    wb = load_workbook(YANWEN, read_only=True, data_only=True)
    au_ranges = load_au_zones(wb)
    ca_map = load_ca_fsa_map(wb)
    wb.close()

    doc = {
        "version": 1,
        "sources": {
            "yanwenWorkbook": str(YANWEN),
            "auSheet": "国家分组-澳洲邮编",
            "caSheet": "加拿大专线产品邮编分区",
            "cainiaoAuNote": "Bundled Cainiao XLSX has no AU postcode sheet; 澳大利亚/N区 uses the same zone digits as this Yanwen AU extract.",
        },
        "au": {"numericRanges": au_ranges},
        "ca": {"postalKeysToZone": ca_map},
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(doc, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print("Wrote", OUT)
    print("  AU intervals:", len(au_ranges), " CA keys:", len(ca_map))


if __name__ == "__main__":
    main()
