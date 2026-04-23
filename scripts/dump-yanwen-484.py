# -*- coding: utf-8 -*-
import sys
from pathlib import Path
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

p = Path(r"c:\Users\Administrator\Desktop\东莞燕文报价单20260413版.xlsx")
wb = load_workbook(p, read_only=True, data_only=True)

for name in wb.sheetnames:
    if "专线追踪" in name and "特货" in name and "专属" not in name and "Coupang" not in name and "AE" not in name and "Temu" not in name:
        print("===", name, "===")
        ws = wb[name]
        for i, row in enumerate(ws.iter_rows(max_row=45, values_only=True)):
            print(i, row)

wb.close()
