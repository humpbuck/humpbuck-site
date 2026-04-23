"""One-off: print structure of logistics xlsx files."""
import sys
from pathlib import Path

from openpyxl import load_workbook

paths = [
    Path(r"c:\Users\Administrator\Desktop\《菜鸟跨境产品报价表-20260418010526》.xlsx"),
    Path(r"c:\Users\Administrator\Desktop\东莞燕文报价单20260413版.xlsx"),
]

for p in paths:
    print("=" * 80)
    print(p.name)
    if not p.exists():
        print("MISSING")
        continue
    wb = load_workbook(p, read_only=True, data_only=True)
    for sn in wb.sheetnames[:25]:
        ws = wb[sn]
        rows = []
        for i, row in enumerate(ws.iter_rows(max_row=40, values_only=True)):
            rows.append(row)
        print(f"\n--- Sheet: {sn!r} ({ws.max_row} rows) ---")
        for i, row in enumerate(rows[:25]):
            print(i, row)
    wb.close()
