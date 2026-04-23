# -*- coding: utf-8 -*-
"""Export Cainiao S5059/OH + Yanwen 484 rows to lib/data/logistics-rates.json."""
import json
import re
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "lib" / "data" / "logistics-rates.json"

CAINIAO = Path(r"c:\Users\Administrator\Desktop\《菜鸟跨境产品报价表-20260418010526》.xlsx")
YANWEN = Path(r"c:\Users\Administrator\Desktop\东莞燕文报价单20260413版.xlsx")


def parse_weight_band(s: str) -> tuple[float, float] | None:
    if not s or not isinstance(s, str):
        return None
    s = s.strip().replace("＜", "<").replace("≤", "<=").replace(" ", "")
    # patterns: 0<W≤0.1, 0.1<W≤0.2, 0<W≤1
    m = re.match(r"^([\d.]+)<W<=([\d.]+)$", s)
    if m:
        return float(m.group(1)), float(m.group(2))
    m = re.match(r"^([\d.]+)<W≤([\d.]+)$", s)
    if m:
        return float(m.group(1)), float(m.group(2))
    return None


def load_cainiao_sheet(wb, sheet_name: str) -> dict[str, list[dict]]:
    """country name (Chinese) -> bands with wMin, wMax, rmbPerKg, regRmb."""
    ws = wb[sheet_name]
    by_country: dict[str, list[dict]] = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        country = row[1]
        band_s = row[4]
        rmb_kg = row[5]
        reg = row[6]
        if not country or not band_s:
            continue
        if not isinstance(country, str):
            continue
        band = parse_weight_band(str(band_s))
        if not band:
            continue
        try:
            pk = float(rmb_kg)
            pr = float(reg)
        except (TypeError, ValueError):
            continue
        c = country.strip()
        by_country.setdefault(c, []).append(
            {"wMin": band[0], "wMax": band[1], "rmbPerKg": pk, "regRmb": pr}
        )
    for c in by_country:
        by_country[c].sort(key=lambda x: (x["wMin"], x["wMax"]))
    return by_country


def load_yanwen_484(
    wb, name_to_iso_hint: dict[str, str] | None = None
) -> tuple[dict[str, list | dict], dict[str, str]]:
    """Parse 燕文专线追踪-特货. Canada / Australia → nested zone maps 1–4; others flat lists."""
    ws = wb["燕文专线追踪-特货"]
    iso_map: dict[str, str] = dict(name_to_iso_hint or {})
    by_iso: dict[str, list | dict] = {}
    zone_split = {"加拿大", "澳大利亚"}

    def append_band(iso: str, band: dict, zone_digit: str | None) -> None:
        if zone_digit is not None:
            cur = by_iso.get(iso)
            if not isinstance(cur, dict):
                cur = {}
                by_iso[iso] = cur
            cur.setdefault(zone_digit, []).append(band)
            return
        cur = by_iso.get(iso)
        if isinstance(cur, dict):
            return
        if cur is None:
            by_iso[iso] = []
        assert isinstance(by_iso[iso], list)
        by_iso[iso].append(band)

    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or row[0] is None:
            continue
        name = row[1]
        code = row[2]
        if not isinstance(name, str) or name.strip() in ("国家", "Country"):
            continue
        name_k = name.strip()
        z = str(code).strip() if code is not None else ""
        zone_digit: str | None = None
        if name_k in zone_split:
            m = re.match(r"^([1-4])区$", z)
            if not m:
                continue
            zone_digit = m.group(1)
        iso = None
        if isinstance(code, str):
            cc = code.strip().upper()
            if len(cc) == 2 and cc.isalpha():
                iso = cc
        if iso is None:
            iso = iso_map.get(name_k)
        if iso is None:
            continue
        if name_k not in iso_map:
            iso_map[name_k] = iso
        try:
            pk = float(row[3])
            fee = float(row[4])
        except (TypeError, ValueError):
            continue
        band_s = row[5]
        min_w = row[6]
        wmin, wmax = None, None
        if isinstance(band_s, str) and " - " in band_s:
            a, b = band_s.split(" - ", 1)
            try:
                wmin = float(a.strip())
                wmax = float(b.strip())
            except ValueError:
                pass
        if wmin is None or wmax is None:
            continue
        min_charge = None
        if min_w is not None:
            try:
                min_charge = float(min_w)
            except (TypeError, ValueError):
                pass
        band = {
            "wMin": wmin,
            "wMax": wmax,
            "rmbPerKg": pk,
            "pieceRmb": fee,
            "minChargeKg": min_charge,
        }
        append_band(iso, band, zone_digit)

    for cc, val in list(by_iso.items()):
        if isinstance(val, list):
            val.sort(key=lambda x: (x["wMin"] or 0, x["wMax"] or 0))
        elif isinstance(val, dict):
            for zd in val:
                val[zd].sort(key=lambda x: (x["wMin"] or 0, x["wMax"] or 0))
    return by_iso, iso_map


def main():
    out: dict = {
        "cainiaoVolumetricDivisor": 8000,
        "yanwenVolumetricDivisor": 18000,
        "yanwenDomesticToWarehouseCny": 5,
        "freeInternationalLegCny": 50,
        "cainiaoPreferenceMarginCny": 5,
        "cainiao": {},
        "yanwen484": {},
        "cainiaoCountryToIso": {},
    }

    prev_iso_hint: dict[str, str] = {}
    prev_fallback = None
    if OUT.is_file():
        try:
            prev_full = json.loads(OUT.read_text(encoding="utf-8"))
            if isinstance(prev_full, dict):
                prev_iso_hint = dict(prev_full.get("cainiaoCountryToIso") or {})
                prev_fallback = prev_full.get("cainiaoIsoFallback")
        except OSError:
            pass

    if CAINIAO.is_file():
        wb = load_workbook(CAINIAO, read_only=True, data_only=True)
        s5059 = load_cainiao_sheet(wb, "菜鸟轻小件-带电（S5059）")
        oh = load_cainiao_sheet(wb, "菜鸟标准挂号-带电（OH）")
        wb.close()
        out["cainiao"]["S5059"] = s5059
        out["cainiao"]["OH"] = oh
    else:
        print("skip Cainiao:", CAINIAO)

    if YANWEN.is_file():
        wb = load_workbook(YANWEN, read_only=True, data_only=True)
        y484, iso_map = load_yanwen_484(wb, prev_iso_hint)
        wb.close()
        out["yanwen484"] = y484
        out["cainiaoCountryToIso"] = iso_map
    else:
        print("skip Yanwen:", YANWEN)
        out["cainiaoCountryToIso"] = prev_iso_hint

    if prev_fallback is not None:
        out["cainiaoIsoFallback"] = prev_fallback

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    print("Wrote", OUT)


if __name__ == "__main__":
    main()
