import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

from openpyxl import load_workbook

paths = [
    Path(r"c:\Users\Administrator\Desktop\《菜鸟跨境产品报价表-20260418010526》.xlsx"),
    Path(r"c:\Users\Administrator\Desktop\东莞燕文报价单20260413版.xlsx"),
]
out = Path("h:/MY-STORES/humpbuck-site/scripts/_xlsx-dump.txt")
lines = []

for p in paths:
    lines.append("=" * 80)
    lines.append(str(p.name))
    if not p.exists():
        lines.append("MISSING")
        continue
    wb = load_workbook(p, read_only=True, data_only=True)
    lines.append("SHEETS: " + ", ".join(wb.sheetnames))
    for sn in wb.sheetnames:
        if any(
            x in sn
            for x in ("S5059", "5059", "OH", "484", "燕文", "专线", "特货")
        ):
            ws = wb[sn]
            lines.append(f"\n>>> SHEET {sn} max_row={ws.max_row}")
            for i, row in enumerate(
                ws.iter_rows(max_row=min(80, ws.max_row or 80), values_only=True)
            ):
                lines.append(f"{i}\t{row}")
    wb.close()

out.write_text("\n".join(lines), encoding="utf-8")
print("Wrote", out)
