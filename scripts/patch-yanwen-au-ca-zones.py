# -*- coding: utf-8 -*-
"""Embed Yanwen 484 AU/CA 1–4区 into lib/data/logistics-rates.json (nested under yanwen484)."""
import json
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
JSON_PATH = ROOT / "lib" / "data" / "logistics-rates.json"
XLSX = Path(r"c:\Users\Administrator\Desktop\东莞燕文报价单20260413版.xlsx")


def rows_for(ws, name_cn: str, zone: str) -> list[dict]:
    out: list[dict] = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or row[0] is None:
            continue
        n, z = row[1], row[2]
        if not isinstance(n, str) or not isinstance(z, str):
            continue
        if n.strip() != name_cn or z.strip() != zone:
            continue
        try:
            pk, fee = float(row[3]), float(row[4])
        except (TypeError, ValueError):
            continue
        band_s = row[5]
        if not isinstance(band_s, str) or " - " not in band_s:
            continue
        a, b = band_s.split(" - ", 1)
        try:
            wmin, wmax = float(a.strip()), float(b.strip())
        except ValueError:
            continue
        min_charge = None
        if row[6] is not None:
            try:
                min_charge = float(row[6])
            except (TypeError, ValueError):
                pass
        out.append(
            {
                "wMin": wmin,
                "wMax": wmax,
                "rmbPerKg": pk,
                "pieceRmb": fee,
                "minChargeKg": min_charge,
            }
        )
    out.sort(key=lambda x: (x["wMin"], x["wMax"]))
    return out


def main() -> None:
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb["燕文专线追踪-特货"]
    zoned: dict[str, dict[str, list[dict]]] = {}
    for iso, cn in (("CA", "加拿大"), ("AU", "澳大利亚")):
        zoned[iso] = {}
        for zi in ("1", "2", "3", "4"):
            zoned[iso][zi] = rows_for(ws, cn, f"{zi}区")
    wb.close()

    data = json.loads(JSON_PATH.read_text(encoding="utf-8"))
    y = data.setdefault("yanwen484", {})
    for iso, zones in zoned.items():
        y[iso] = zones
    JSON_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print("Patched yanwen484", list(zoned.keys()), "with zones 1–4.")


if __name__ == "__main__":
    main()
